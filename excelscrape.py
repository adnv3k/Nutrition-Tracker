"""
This extracts data from the table
A1-2 on pg. 133 of the Dietary_Guidelines_for_Americans-2020-2025.pdf,
and saves it to a shelf, "usda".
A function to get the specific dietary 
recommendation for a given sex and age is implemented.
"""

import os
from re import split
import shelve
import openpyxl

# Reading excel
#
os.chdir(r'.\\USDA') # Change dir to location of .xlsx
workbook = openpyxl.load_workbook("Daily Nutritional Goals.xlsx")
sheet = workbook.copy_worksheet(workbook.worksheets[0])
os.chdir(r'..\\')
# Checking results
#
# print(type(sheet))
# print(sheet.columns)
#
# print(workbook.sheetnames)
# print(sheet['A4'])
# cell = sheet['C2']
# print(type(cell.value))
# print(cell.value)
#
# sheet.cell(row=1, column=2)
# sheet.cell(row=1, column=2).value

# Grabbing nutrient names
#
# macronutrients = []
# for i in range(5,15):
#     macronutrients.append(str(sheet.cell(row=i,column=1).value))
# # print(macronutrients)
#
# minerals = []
# for i in range(16,23):
#     minerals.append(str(sheet.cell(row=i,column=1).value))
# # print(minerals)
#
# vitamins = []
# for i in range(24,36):
#     vitamins.append(str(sheet.cell(row=i,column=1).value))
# # print(vitamins)
#
# Save nutrient names
# daily = {"macronutrients": macronutrients, "minerals": minerals, "vitamins": vitamins}
# os.chdir(r'..\\')
# file = shelve.open('usda')
# file['daily_nutrition'] = daily
# file.close()

# Associate nutrient names with a value
#
# print(sheet.cell(row=2,column=13).value)
# column = 3
#
# test = sheet.cell(row=2,column=column).value
# print(f'Sex: {test[0]}')
# for ele in test:
#     print(ele)
# print(len(test))
# test = test.split('\n')
# print(test)
# age_range = test[1].split('-')
# print(age_range)
# f=shelve.open[3] # Used to return error to stop code

"""
The idea is to input your sex and age, 
then it will return the appropriate daily nutrition recommendation.
Save every group:
group[sex][age_group] = value
Structure:
daily (contains categorized data from the table)
    sex
        age_group
Example of input
input: 
sex = Female
input: age = 26
for age_group in [*daily['F']]:
    if age <= upper_age_range:
        load this group
"""
# Extract from excel table
def extract_table():
    daily = {"F": {}, "M": {}}
    for column in range(3,16):
        # Grab nutrient names and values
        calories = {'Calories': sheet.cell(row=3,column=column).value}
        # print(calories)

        macronutrients = {}
        for i in range(5,15):
            macronutrients[str(sheet.cell(row=i,column=1).value)] = sheet.cell(row=i,column=column).value
        # print(macronutrients)

        minerals = {}
        for i in range(16,23):
            minerals[str(sheet.cell(row=i,column=1).value)] = sheet.cell(row=i,column=column).value
        # print(minerals)

        vitamins = {}
        for i in range(24,36):
            vitamins[str(sheet.cell(row=i,column=1).value)] = sheet.cell(row=i,column=column).value
        # print(vitamins)
        # Put nutrient names and values in a dict
        values = {"calories": calories, "macronutrients": macronutrients, "minerals": minerals, "vitamins": vitamins}
        # Get sex and age groups
        group = sheet.cell(row=2,column=column).value.split('\n')
        if group[0] == "M/F": # Edge case
            upper_range = int(group[1].split('-')[-1])
            lower_range = int(group[1].split('-')[0])
            daily['F'][(lower_range,upper_range)] = values
            daily['M'][(lower_range,upper_range)] = values
            continue
        if "+" in group[1]: # Edge case
            upper_range = 130
            lower_range = 51
        else:
            upper_range = int(group[1].split('-')[-1])
            lower_range = int(group[1].split('-')[0])
        daily[group[0]][(lower_range,upper_range)] = values
    return daily

daily = extract_table()
# print(daily)

# Save daily to shelf
# file = shelve.open('usda')
# file['daily_nutrition'] = daily
# file.close()
#
# Load daily from shelf
file = shelve.open('usda')
daily = file['daily_nutrition']
file.close()

# Get daily nutritional recommendation for age and sex
def get_daily_nutrition(sex: str, age: int):
    # Converts sex to correct format
    sex = sex[0].upper()

    for age_range in [*daily[sex]]:
        if age in range(age_range[0],age_range[1]+1):
            return daily[sex][age_range]

daily_nutrition = get_daily_nutrition(age=35, sex="male")
# print(daily_nutrition)

# Conversion rates for values in table
#
# 1 IU retinol = 0.3 mcg RAE
#
# 18:2 lineoic acid 
# PUFA 18:2 n-6 c,c
# PUFA 18:2
# Fatty acids, total polyunsaturated
#
# 18:3 lineoic acid 
# PUFA 18:3 n-3 c,c,c (ALA)
# PUFA 18:3
# Fatty acids, total polyunsaturated